#PSEUDOCODE
#import tkinter 
from tkinter import*
from tkinter import ttk
import tkinter
import os
import openpyxl




#create class
class Tracer:
    #init function & create instance of variable
    def __init__(self):
            
            #create tkinter window   
            self.ws = tkinter.Tk()
            self.ws.geometry("1000x600")
            self.ws.title("COVID Contact Tracing Form")
            self.ws.configure(highlightbackground= "dark blue", highlightthickness= 6)

        #create frame for necessary information 
            self.user_info_frame = tkinter.Frame(self.ws, background= "light blue", highlightthickness= 4, highlightbackground= "dark blue")
            self.user_info_frame.place(x= 10, y = 20, width = 480, height = 540)
            #label the frame
            self.user_info_frame_label =tkinter.LabelFrame(self.user_info_frame, text="PERSONAL INFORMATION",
                                                           padx=5, pady= 20, font= ("Brass Mono", 15),
                                                            highlightthickness= 4, highlightbackground= "light yellow", 
                                                            foreground= "dark blue", highlightcolor= "dark green")
            self.user_info_frame_label.place(x= 10, y = 20, width= 460, height= 500)

        #create frame for checkboxes
            self.symptoms_frame= tkinter.Frame(self.ws, background= "light blue",highlightthickness= 4, highlightbackground= "dark blue")
            self.symptoms_frame.place (x= 500, y = 20, width = 480, height = 540)
            #label the frame
            self.symptoms_frame_label =tkinter.LabelFrame(self.symptoms_frame, text=  "SYMPTOMS ", 
                                                          padx=5, pady= 20, font= ("Brass Mono", 15),
                                                          foreground= "dark blue", highlightbackground ="light yellow", 
                                                          highlightthickness= 4, highlightcolor= "dark green")
            self.symptoms_frame_label.place(x= 10, y = 20, width= 460, height= 500)
            
    #create functions for getting user's input using tkinter entry with label

        #first name label and entry
    def firstName(self):
        self.FIRST_NAME = tkinter.Label(self.user_info_frame_label, text="First Name")
        self.FIRST_NAME.grid(row=0, column=0, padx= 10, pady= 5)
        self.FIRST_NAME_INPUT = tkinter.Entry(self.user_info_frame_label)
        self.FIRST_NAME_INPUT.grid(row=0, column=1)
        
        #last name label and entry
        def lastName(self):
            self.LAST_NAME = tkinter.Label(self.user_info_frame, text="Last Name", background= 'light yellow', highlightbackground= "dark blue", highlightthickness= 2)
            self.LAST_NAME.place(x = 20, y = 20)
            self.LAST_NAME_INPUT = tkinter.Entry(self.user_info_frame)
            self.LAST_NAME_INPUT.place(x = 25, y = 40)

        #Gender label and entry
        def gender(self):
            self.GENDER = tkinter.Label(self.user_info_frame, text="Gender")
            self.GENDER.grid(row=2, column=0)
            self.GENDER_INPUT = ttk.Combobox(self.user_info_frame, values=["Male", "Female", "Transgender Male", "Transgender Female", "Prefer Not to Say"])
            self.GENDER_INPUT.grid(row=2, column=1)

        
        #Age label and entry
        def age(self):
            self.AGE = tkinter.Label(self.user_info_frame, text="Age")
            self.AGE.grid(row=3, column=0)
            self.AGE_INPUT = tkinter.Spinbox(self.user_info_frame_label, from_=1, to=100)
            self.AGE_INPUT.grid(row=3, column=1)

        #Nationality label and entry
        def national(self):
            #Nationality label and entry
            self.NATIONALITY = tkinter.Label(self.user_info_frame, text="Nationality")
            self.NATIONALITY.grid(row=4, column=0)
            self.NATIONALITY_INPUT = ttk.Combobox(self.user_info_frame, values=["Filipino", "Chinese", "Korean", "American", ])
            self.NATIONALITY_INPUT['state'] = 'normal'
            self.NATIONALITY_INPUT.grid(row=4, column=1)
            
        #Vaccination Status label and entry
        def vaccination(self):
            #"Unvaccinated", "1st Dose", "2nd Dose(Fully Vaccinated)", "1st Booster Shot", "2nd Booster Shot"
            self.VACCINATION_STATUS = tkinter.Label(self.symptoms_frame, text="Vaccination Status")
            self.VACCINATION_STATUS.grid(row=7, column=0)
            self.VACCINATION_STATUS_INPUT = ttk.Combobox(self.symptoms_frame, values= ["Unvaccinated", "1st Dose", "2nd Dose(Fully Vaccinated)", "1st Booster Shot", "2nd Booster Shot"])
            self.VACCINATION_STATUS_INPUT['state'] = 'readonly'
            self.VACCINATION_STATUS_INPUT.grid(row=7, column=1)

        #Phone Number label and entry
        def phone(self):
            self.PHONE = tkinter.Label(self.user_info_frame_label, text= "Cellphone Number")
            self.PHONE.grid(row=6, column=0)
            self.PHONE_INPUT = tkinter.Entry(self.user_info_frame_label)
            self.PHONE_INPUT.grid(row=6, column=1)
        
        # address label and entry
        def address(self):
            self.ADDRESS = tkinter.Label(self.user_info_frame_label, text="Address")
            self.ADDRESS.grid(row=7, column=0)
            self.ADDRESS_INPUT = tkinter.Entry(self.user_info_frame_label)
            self.ADDRESS_INPUT.grid(row=7, column=1)    
        #put terms and condition checkbox
        def accept(self):
            self.accept_var = tkinter.StringVar(value="Not Accepted")
            self.terms_check = tkinter.Checkbutton(self.user_info_frame_label, text= "I accept the terms and conditions.",
                                            variable=self.accept_var, onvalue="Accepted", offvalue="Not Accepted")
            self.terms_check.grid(row=8, column=1)
        #fix spacing of widgets
        def separate(self):
            for widget in self.user_info_frame_label.winfo_children():
                widget.grid_configure(padx=20, pady=10)

        #create a save/add function for all user's info into a database ( i prefer excel as my database)
        def add(self):
            accepted = self.accept_var.get()
            if accepted=="Accepted":
                firstname = self.FIRST_NAME_INPUT.get()
                lastname = self.LAST_NAME_INPUT.get()
                nationality  = (self.NATIONALITY_INPUT).get() 
                age = self.AGE_INPUT.get()
                address = self.ADDRESS_INPUT.get()
                gender = self.GENDER_INPUT.get()
                vaccination = self.VACCINATION_STATUS_INPUT.get()
                phone = self.PHONE_INPUT.get()
                if firstname and lastname and age and gender and nationality and address and phone and vaccination:
                        print("First name: ", firstname , "\n" , "Last name: ", lastname)
                        print("Age: ", age, "\n", "Nationality: ", nationality, "\n", "Gender :", gender)
                        print("Vaccination Status: ", vaccination, "\n", "Cellphone Number : ", phone, "\n", "Address :", address)

                        filepath =("C:\git\BSCPE 1ST YEAR 2ND SEM\OOP\ASSIGNMENTS\COVID-CONTACT-TRACING-APP\Tracer\data.xlsx")

                        if not os.path.exists(filepath):
                            workbook = openpyxl.Workbook()
                            sheet = workbook.active
                            heading = ["First Name", "Last Name", "Gender", "Age", "Nationality",
                                    "Address", "Cellphone Number", "Vaccination status"]
                            sheet.append(heading)
                            workbook.save(filepath)
                        workbook = openpyxl.load_workbook(filepath)
                        sheet = workbook.active
                        sheet.append([firstname, lastname, gender, age, nationality, address,
                                    phone, vaccination])
                        workbook.save(filepath)
                    
                else:
                            tkinter.messagebox.showwarning(title="Error", message="All information are required.")
            else:
                    tkinter.messagebox.showwarning(title= "Error", message="You have not accepted the terms and conditions")
        #create search method
        def searchByFirstName(self):
            FirstNameInput = self.FIRST_NAME_INPUT.get()
            self.FIRST_NAME_INPUT.configure(state= tkinter.NORMAL)
            self.LAST_NAME_INPUT.configure(state= tkinter.NORMAL)
            self.GENDER_INPUT.configure(state= tkinter.NORMAL)
            self.AGE_INPUT.configure(state= tkinter.NORMAL)
            self.NATIONALITY_INPUT.configure(state= tkinter.NORMAL)
            self.ADDRESS_INPUT.configure(state= tkinter.NORMAL)
            self.PHONE_INPUT.configure(state= tkinter.NORMAL)
            self.VACCINATION_STATUS_INPUT.configure(state= tkinter.NORMAL)
            
            self.FIRST_NAME_INPUT.delete(0, 'end')
            self.LAST_NAME_INPUT.delete(0, 'end')
            self.GENDER_INPUT.delete(0, 'end')
            self.AGE_INPUT.delete(0, 'end')
            self.NATIONALITY_INPUT.delete(0, 'end')
            self.ADDRESS_INPUT.delete(0, 'end')
            self.PHONE_INPUT.delete(0, 'end')
            self.VACCINATION_STATUS_INPUT.delete(0, 'end')
                
            wb = openpyxl.load_workbook("C:\git\BSCPE 1ST YEAR 2ND SEM\OOP\ASSIGNMENTS\COVID-CONTACT-TRACING-APP\Tracer\data.xlsx")  
            sheet = wb["Sheet"]  
            for cell in sheet.iter_rows(min_row=1, min_col=1, max_row =sheet.max_row, max_col=9, values_only=True):  
                if cell[0] == str(FirstNameInput):
                    self.FIRST_NAME_INPUT.insert(0, cell[0])
                    self.LAST_NAME_INPUT.insert(0, cell[1])
                    self.GENDER_INPUT.insert(0, cell[2])
                    self.AGE_INPUT.insert(0, cell[3])
                    self.NATIONALITY_INPUT.insert(0, cell[4])
                    self.ADDRESS_INPUT.insert(0, cell[5])
                    self.PHONE_INPUT.insert(0, cell[6])
                    self.VACCINATION_STATUS_INPUT.insert(0, cell[7])
                    
                    
                    self.FIRST_NAME_INPUT.configure(state= tkinter.DISABLED)
                    self.LAST_NAME_INPUT.configure(state= tkinter.DISABLED)
                    self.GENDER_INPUT.configure(state= tkinter.DISABLED)
                    self.AGE_INPUT.configure(state= tkinter.DISABLED)
                    self.NATIONALITY_INPUT.configure(state= tkinter.DISABLED)
                    self.ADDRESS_INPUT.configure(state= tkinter.DISABLED)
                    self.PHONE_INPUT.configure(state= tkinter.DISABLED)
                    self.VACCINATION_STATUS_INPUT.configure(state= tkinter.DISABLED)
                    
                        
        #create save/add button 
        def button_add(self):
            button = tkinter.Button(self.user_info_frame_label, text="Enter data", command= self.add)
            button.grid(row=9, column=2, sticky="news", padx=20, pady=10)    
                
        #create search button
        def button_search(self):
            button = tkinter.Button(self.user_info_frame_label, text="Search data", command= self.searchByFirstName)
            button.grid(row=9, column=3, sticky="news", padx=20, pady=10)
            